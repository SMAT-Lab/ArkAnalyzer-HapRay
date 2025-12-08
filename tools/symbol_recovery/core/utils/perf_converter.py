#!/usr/bin/env python3
"""
从perf.data 或 excel 文件读取缺失符号的函数，进行反汇编和LLM分析
"""

import os
import shutil
import sqlite3
from collections import defaultdict
from pathlib import Path
from typing import Optional

import pandas as pd
from elftools.elf.elffile import ELFFile

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - optional dependency
    Workbook = None
    Alignment = None
    get_column_letter = None

try:
    from core.llm.batch_analyzer import BatchLLMFunctionAnalyzer
except ImportError:  # pragma: no cover - optional dependency
    BatchLLMFunctionAnalyzer = None
from core.llm.initializer import init_llm_analyzer
from core.utils import StringExtractor, get_logger
from core.utils import common as util
from core.utils.config import (
    CALL_COUNT_ANALYSIS_PATTERN,
    DEFAULT_BATCH_SIZE,
    DEFAULT_LLM_MODEL,
    DEFAULT_TOP_N,
    config,
)

logger = get_logger(__name__)

# HAP 地址解析（延迟导入，避免循环依赖）
try:
    from core.utils.hap_address_resolver import (
        is_hap_address,
        resolve_hap_address_from_perfdb,
        resolve_hap_addresses_batch,
    )

    HAP_RESOLVER_AVAILABLE = True
except ImportError:
    HAP_RESOLVER_AVAILABLE = False
    is_hap_address = None
    resolve_hap_address_from_perfdb = None
    resolve_hap_addresses_batch = None
    logger.warning('HAP address resolver module unavailable')

try:
    import r2pipe

    from core.analyzers.r2_analyzer import R2FunctionAnalyzer

    R2_AVAILABLE = True
except ImportError:
    R2_AVAILABLE = False
    R2FunctionAnalyzer = None
    r2pipe = None
    logger.warning('r2_function_analyzer module unavailable, will use capstone for disassembly and LLM analysis')


def _check_r2_actually_available():
    """
    检测 radare2 是否真的可用（不仅仅是模块是否可导入）
    通过检查 r2 命令是否在 PATH 中来判断

    Returns:
        bool: 如果 radare2 可用返回 True，否则返回 False
    """
    if not R2_AVAILABLE:
        return False

    try:
        # 检查 r2 命令是否在 PATH 中
        r2_path = shutil.which('r2')
        if r2_path:
            logger.debug(f'Found radare2 command: {r2_path}')
            return True

        # 如果找不到 r2 命令，说明 radare2 不可用
        logger.debug('radare2 command not found (r2 not in PATH)')
        return False
    except Exception as e:
        logger.debug(f'Error checking radare2 availability: {e}')
        return False


class MissingSymbolFunctionAnalyzer:
    """分析缺失符号的函数"""

    def __init__(
        self,
        excel_file=None,
        so_dir=None,
        perf_db_file=None,
        use_llm=True,
        llm_model=None,
        batch_size=None,
        context=None,
        use_capstone_only=False,
        save_prompts=False,
        output_dir=None,
        skip_decompilation=False,
        open_source_lib=None,
    ):
        """
        初始化分析器

        Args:
            excel_file: 缺失符号分析 Excel 文件（可选，如果提供 perf_db_file 则不需要）
            so_dir: SO 文件目录
            perf_db_file: perf.db 文件路径（可选，如果提供则直接从数据库读取，不需要 Excel）
            use_llm: 是否使用 LLM 分析
            llm_model: LLM 模型名称
            use_batch_llm: 是否使用批量 LLM 分析（一个 prompt 包含多个函数，默认 True）
            batch_size: 批量分析时每个 prompt 包含的函数数量（默认 3）
            context: 自定义上下文信息（可选，如果不提供则根据 SO 文件名自动推断）
            use_capstone_only: 只使用 Capstone 反汇编（不使用 Radare2，即使已安装）
            save_prompts: 是否保存生成的 prompt 到文件
            output_dir: 输出目录，用于保存 prompt 文件
            skip_decompilation: 是否跳过反编译（默认 False，启用反编译可提高 LLM 分析质量但较慢）
        """
        self.excel_file = Path(excel_file) if excel_file else None
        self.perf_db_file = Path(perf_db_file) if perf_db_file else None
        self.so_dir = Path(so_dir) if so_dir else None
        self.use_llm = use_llm
        self.llm_model = llm_model if llm_model is not None else DEFAULT_LLM_MODEL
        self.batch_size = batch_size if batch_size is not None else DEFAULT_BATCH_SIZE
        self.context = context  # 自定义上下文
        self.use_capstone_only = use_capstone_only  # 强制使用 Capstone
        self.skip_decompilation = skip_decompilation  # 是否跳过反编译
        self._r2_actually_available = None  # 缓存 radare2 实际可用性（延迟检测）

        # 验证输入：必须提供 excel_file 或 perf_db_file 之一
        if not self.excel_file and not self.perf_db_file:
            raise ValueError('必须提供 excel_file 或 perf_db_file 之一')

        if self.excel_file and not self.excel_file.exists():
            raise FileNotFoundError(f'Excel 文件不存在: {excel_file}')

        if self.perf_db_file and not self.perf_db_file.exists():
            raise FileNotFoundError(f'perf.db 文件不存在: {perf_db_file}')

        if self.so_dir:
            # 如果 so_dir 是文件（通过 --so-file 指定），允许；如果是目录，检查是否存在
            if not self.so_dir.exists():
                raise FileNotFoundError(f'SO 文件或目录不存在: {so_dir}')
            if self.so_dir.is_file() and not self.so_dir.suffix == '.so':
                raise ValueError(f'指定的文件不是 SO 文件: {so_dir}')

        # 初始化反汇编器
        self.md = util.create_disassembler()

        # 初始化字符串提取器（稍后设置 disassemble_func）
        self.string_extractor = None

        # 缓存 radare2 分析器实例（按 SO 文件路径缓存，避免重复初始化）
        self._r2_analyzers = {}  # {so_file_path: R2FunctionAnalyzer}

        # 初始化 LLM 分析器（使用公共工具函数，避免重复代码）
        self.llm_analyzer, self.use_llm, self.use_batch_llm = init_llm_analyzer(
            use_llm=self.use_llm,
            llm_model=self.llm_model,
            batch_size=self.batch_size,
            save_prompts=save_prompts,
            output_dir=output_dir,
            open_source_lib=open_source_lib,
        )

    def find_so_file(self, file_path):
        """
        根据文件路径找到对应的 SO 文件

        Args:
            file_path: 文件路径（如 /proc/.../libxwebcore.so）

        Returns:
            SO 文件路径，如果找不到返回 None
        """
        # 如果 so_dir 是一个文件（通过 --so-file 指定），直接返回它（如果文件名匹配）
        if self.so_dir and self.so_dir.is_file():
            # 提取目标文件名（从 file_path）
            target_so_name = file_path.split('/')[-1] if '/' in file_path else file_path
            # 如果指定的 SO 文件名匹配，直接返回
            if self.so_dir.name == target_so_name:
                return self.so_dir.resolve()
            # 如果不匹配，返回 None（只分析指定的 SO 文件）
            # 注意：这里不匹配可能是因为 file_path 来自调用栈，而不是函数所在的文件
            # 真正的 SO 文件名可能在 address 中，但这里我们无法访问 address
            # 所以这个检查会在 analyze_function 调用时再次进行
            return None

        # 提取文件名（如 libxwebcore.so）
        so_name = None
        so_name = file_path.split('/')[-1] if '/' in file_path else file_path

        # 在指定目录中查找
        so_file = self.so_dir / so_name
        if so_file.exists():
            # 返回绝对路径，确保路径一致性（用于缓存键）
            return so_file.resolve()

        # 如果找不到，尝试递归查找所有子目录中的 SO 文件
        for so_file in self.so_dir.rglob('*.so'):
            if so_file.name == so_name:
                # 返回绝对路径，确保路径一致性（用于缓存键）
                return so_file.resolve()

        return None

    def extract_offset_from_address(self, address):
        """
        从地址字符串中提取偏移量（虚拟地址）
        使用统一的 parse_offset 函数
        """
        return util.parse_offset(address)

    def vaddr_to_file_offset(self, elf_file, vaddr):
        """将虚拟地址转换为文件偏移量"""
        for segment in elf_file.iter_segments():
            if segment['p_type'] == 'PT_LOAD':
                vaddr_start = segment['p_vaddr']
                vaddr_end = vaddr_start + segment['p_memsz']
                if vaddr_start <= vaddr < vaddr_end:
                    return segment['p_offset'] + (vaddr - vaddr_start)
        return None

    def find_function_start(self, elf_file, vaddr):
        """查找函数的起始位置（使用统一的 util.find_function_start）"""
        return util.find_function_start(elf_file, vaddr, self.md)

    def disassemble_function(self, elf_file, vaddr, size=2000):
        """反汇编指定虚拟地址的函数代码

        Args:
            elf_file: ELF 文件对象
            vaddr: 函数虚拟地址
            size: 最大反汇编大小（默认 2000 字节，增加以支持大型函数）
        """
        try:
            # 获取 .text 段
            text_section = None
            for section in elf_file.iter_sections():
                if section.name == '.text':
                    text_section = section
                    break

            if not text_section:
                return None

            text_vaddr = text_section['sh_addr']
            text_size = text_section['sh_size']

            # 确保地址在 .text 段内
            if vaddr < text_vaddr or vaddr >= text_vaddr + text_size:
                logger.warning(' Address 0x%s is not in .text section', f'{vaddr:x}')
                return None

            # 查找函数起始位置
            func_start = self.find_function_start(elf_file, vaddr)

            # 尝试从符号表获取函数大小（如果可用）
            func_size = None
            try:
                # 查找 .dynsym 或 .symtab 符号表
                for section_name in ['.dynsym', '.symtab']:
                    symbol_table = elf_file.get_section_by_name(section_name)
                    if symbol_table:
                        for symbol in symbol_table.iter_symbols():
                            if symbol['st_info']['type'] == 'STT_FUNC':
                                sym_addr = symbol['st_value']
                                sym_size = symbol['st_size']
                                # 检查地址是否在这个符号范围内
                                if sym_size > 0 and sym_addr <= vaddr < sym_addr + sym_size:
                                    func_size = sym_size
                                    logger.info(
                                        f'Got function size from symbol table: {func_size} bytes (symbol: {symbol.name})'
                                    )
                                    break
                    if func_size:
                        break
            except Exception:
                # 符号表查找失败，使用默认大小
                pass

            # 计算相对偏移量
            relative_start = func_start - text_vaddr
            if func_size:
                # 使用符号表中的函数大小，但不超过 size 限制
                relative_end = min(relative_start + func_size, relative_start + size, text_size)
            else:
                relative_end = min(relative_start + size, text_size)

            # 读取代码
            code = text_section.data()[relative_start:relative_end]

            if not code:
                return None

            # 反汇编
            instructions = []
            ret_count = 0  # 记录遇到的 ret 指令数量
            consecutive_ret = 0  # 连续 ret 指令计数

            for inst in self.md.disasm(code, func_start):
                instructions.append(inst)

                # 改进的停止条件：
                # 1. 如果遇到 ret 指令，记录但不立即停止
                # 2. 如果连续遇到多个 ret（可能是函数结束标记），且已反汇编足够指令，则停止
                # 3. 如果遇到 ret 且已经反汇编了大部分函数（超过 80%），则停止
                if inst.mnemonic == 'ret':
                    ret_count += 1
                    consecutive_ret += 1
                    # 如果连续遇到 2 个 ret，且已反汇编超过 300 条指令（约 1200 字节），可能是函数结束
                    # 注意：ARM64 指令通常是 4 字节，300 条指令 ≈ 1200 字节
                    # 这个阈值适用于大多数函数，但复杂函数可能超过这个值
                    if consecutive_ret >= 2 and len(instructions) > 300:
                        break
                    # 如果遇到 ret 且已反汇编超过 size 的 80%，可能是函数结束
                    if len(instructions) * 4 > (relative_end - relative_start) * 0.8:
                        break
                else:
                    consecutive_ret = 0  # 重置连续 ret 计数

                # 如果已经反汇编了足够多的指令（超过 size 限制），停止
                if len(instructions) * 4 > (relative_end - relative_start):
                    break

            return instructions
        except Exception:
            logger.exception('反汇编失败 (vaddr=0x%x)', vaddr)
            return None

    def _init_string_extractor(self):
        """延迟初始化字符串提取器（需要先定义 disassemble_function）"""
        if self.string_extractor is None:
            self.string_extractor = StringExtractor(disassemble_func=self.disassemble_function, md=self.md)

    def _build_context(self, so_file, file_path=None):
        """
        构建 LLM 分析的上下文信息（根据 SO 文件名和应用路径自动推断）

        Args:
            so_file: SO 文件路径
            file_path: 文件路径（可选，用于推断应用类型）

        Returns:
            上下文字符串
        """
        so_name = Path(so_file).name.lower()
        so_file_name = Path(so_file).name

        # 从文件路径推断应用类型
        app_name = None
        if file_path:
            file_path_lower = file_path.lower()
            if 'taobao' in file_path_lower or 'com.taobao' in file_path_lower:
                app_name = '淘宝（Taobao）'
            elif 'wechat' in file_path_lower or 'com.tencent.wechat' in file_path_lower:
                app_name = '微信（WeChat）'
            elif 'alipay' in file_path_lower or 'com.alipay' in file_path_lower:
                app_name = '支付宝（Alipay）'
            elif 'qq' in file_path_lower and 'com.tencent' in file_path_lower:
                app_name = 'QQ'
            elif 'douyin' in file_path_lower or 'com.ss.android' in file_path_lower:
                app_name = '抖音（Douyin）'

        # 根据 SO 文件名推断库的类型和用途
        if 'xwebcore' in so_name or 'xweb' in so_name:
            context = (
                f'这是一个基于 Chromium Embedded Framework (CEF) 的 Web 核心库（{so_file_name}），'
                f'运行在 HarmonyOS 平台上。该库负责网页渲染、网络请求、DOM 操作等核心功能。'
            )
        elif 'wechat' in so_name or 'wx' in so_name:
            context = (
                f'这是一个来自微信（WeChat）应用的 SO 库（{so_file_name}），'
                f'运行在 HarmonyOS 平台上。该库负责即时通讯、社交网络、多媒体处理等功能。'
            )
        elif 'taobao' in so_name or 'tb' in so_name:
            context = (
                f'这是一个来自淘宝（Taobao）应用的 SO 库（{so_file_name}），'
                f'运行在 HarmonyOS 平台上。该库负责电商购物、商品展示、支付处理等功能。'
            )
        elif 'chromium' in so_name or 'blink' in so_name or 'v8' in so_name:
            context = (
                f'这是一个基于 Chromium/Blink 引擎的组件库（{so_file_name}），'
                f'通常用于 Web 渲染、JavaScript 执行等 Web 相关功能。'
            )
        elif 'flutter' in so_name:
            context = (
                f'这是一个 Flutter 框架相关的 SO 库（{so_file_name}），'
                f'Flutter 是 Google 开发的跨平台 UI 框架，用于构建移动应用界面。'
            )
        # 通用格式，根据应用名称调整
        elif app_name:
            context = f'这是一个来自 {app_name} 应用的 SO 库（{so_file_name}），运行在 HarmonyOS 平台上。'
        else:
            context = f'这是一个 SO 库（{so_file_name}），来自 {Path(so_file).parent.name} 目录。'

        return context

    def _get_call_stack_info(self, file_path: str, address: str, vaddr: int) -> Optional[dict]:
        """从 perf.db 获取调用堆栈信息

        Args:
            file_path: 文件路径
            address: 地址字符串
            vaddr: 虚拟地址偏移量

        Returns:
            调用堆栈信息字典，包含调用者和被调用者信息
        """
        if not self.perf_db_file or not self.perf_db_file.exists():
            return None

        try:
            conn = sqlite3.connect(str(self.perf_db_file))
            cursor = conn.cursor()

            try:
                # 加载映射关系
                cursor.execute('SELECT DISTINCT file_id, path FROM perf_files WHERE path IS NOT NULL')
                file_id_to_path = {row[0]: row[1] for row in cursor.fetchall()}

                cursor.execute('SELECT id, data FROM data_dict WHERE data IS NOT NULL')
                name_to_data = {row[0]: row[1] for row in cursor.fetchall()}

                # 查找文件 ID
                file_id = None
                for fid, path in file_id_to_path.items():
                    if path == file_path:
                        file_id = fid
                        break

                if file_id is None:
                    return None

                # 查找地址 ID
                name_id = None
                for nid, data in name_to_data.items():
                    if data == address:
                        name_id = nid
                        break

                if name_id is None:
                    return None

                # 查询调用堆栈：找到调用这个函数的函数（depth 更大的，因为 depth 越大代表源头/调用者）
                # 直接在同一个 callchain_id 中查找，不需要 JOIN perf_sample
                cursor.execute(
                    """
                    SELECT DISTINCT pc2.file_id, pc2.name, pc2.depth, pc2.symbol_id as caller_symbol_id
                    FROM perf_callchain pc1
                    JOIN perf_callchain pc2 ON pc1.callchain_id = pc2.callchain_id
                    WHERE pc1.file_id = ? AND pc1.name = ? AND pc1.symbol_id = -1
                      AND pc2.depth > pc1.depth
                      AND pc2.symbol_id != -1
                    ORDER BY pc2.depth ASC
                    LIMIT 10
                """,
                    (file_id, name_id),
                )

                callers = []
                for row in cursor.fetchall():
                    caller_file_id, caller_name_id, caller_depth, caller_symbol_id = row
                    caller_file_path = file_id_to_path.get(caller_file_id, '')
                    caller_address = name_to_data.get(caller_name_id, '')

                    # 获取调用者函数名（如果有符号）
                    caller_symbol_name = None
                    if caller_symbol_id and caller_symbol_id != -1:
                        try:
                            cursor.execute('SELECT name FROM perf_symbols WHERE id = ?', (caller_symbol_id,))
                            symbol_row = cursor.fetchone()
                            if symbol_row:
                                caller_symbol_name = symbol_row[0]
                        except sqlite3.OperationalError:
                            # perf_symbols 表可能不存在，忽略
                            pass

                    if caller_file_path and caller_address:
                        callers.append(
                            {
                                'file_path': caller_file_path,
                                'address': caller_address,
                                'symbol_name': caller_symbol_name,
                                'depth': caller_depth,
                            }
                        )

                # 查询被调用者：找到这个函数调用的函数（depth 更小的，因为 depth 越小代表被调用到的）
                # 直接在同一个 callchain_id 中查找，不需要 JOIN perf_sample
                cursor.execute(
                    """
                    SELECT DISTINCT pc2.file_id, pc2.name, pc2.depth, pc2.symbol_id as callee_symbol_id
                    FROM perf_callchain pc1
                    JOIN perf_callchain pc2 ON pc1.callchain_id = pc2.callchain_id
                    WHERE pc1.file_id = ? AND pc1.name = ? AND pc1.symbol_id = -1
                      AND pc2.depth < pc1.depth
                      AND pc2.symbol_id != -1
                    ORDER BY pc2.depth DESC
                    LIMIT 10
                """,
                    (file_id, name_id),
                )

                callees = []
                for row in cursor.fetchall():
                    callee_file_id, callee_name_id, callee_depth, callee_symbol_id = row
                    callee_file_path = file_id_to_path.get(callee_file_id, '')
                    callee_address = name_to_data.get(callee_name_id, '')

                    # 获取被调用者函数名（如果有符号）
                    callee_symbol_name = None
                    if callee_symbol_id and callee_symbol_id != -1:
                        try:
                            cursor.execute('SELECT name FROM perf_symbols WHERE id = ?', (callee_symbol_id,))
                            symbol_row = cursor.fetchone()
                            if symbol_row:
                                callee_symbol_name = symbol_row[0]
                        except sqlite3.OperationalError:
                            # perf_symbols 表可能不存在，忽略
                            pass

                    if callee_file_path and callee_address:
                        callees.append(
                            {
                                'file_path': callee_file_path,
                                'address': callee_address,
                                'symbol_name': callee_symbol_name,
                                'depth': callee_depth,
                            }
                        )

                return {
                    'callers': callers[:5],  # 限制最多5个调用者
                    'callees': callees[:5],  # 限制最多5个被调用者
                }

            finally:
                conn.close()

        except Exception as e:
            logger.warning(f'Error getting call stack information: {e}')
            return None

    def _enhance_context_with_call_stack(self, base_context: str, call_stack_info: Optional[dict]) -> str:
        """增强上下文信息，添加调用堆栈信息

        Args:
            base_context: 基础上下文信息
            call_stack_info: 调用堆栈信息

        Returns:
            增强后的上下文信息
        """
        if not call_stack_info:
            return base_context

        context_parts = [base_context]

        # 添加调用者信息
        callers = call_stack_info.get('callers', [])
        if callers:
            context_parts.append('\n调用堆栈信息（谁调用了这个函数）:')
            for i, caller in enumerate(callers[:3], 1):  # 只显示前3个
                caller_info = f'  {i}. '
                if caller.get('symbol_name'):
                    caller_info += f'{caller["symbol_name"]} '
                caller_info += f'({caller["file_path"]} {caller["address"]})'
                context_parts.append(caller_info)

        # 添加被调用者信息（有符号的函数）
        callees = call_stack_info.get('callees', [])
        if callees:
            context_parts.append('\n被调用的函数（这个函数调用了哪些有符号的函数）:')
            for i, callee in enumerate(callees[:3], 1):  # 只显示前3个
                callee_info = f'  {i}. '
                if callee.get('symbol_name'):
                    callee_info += f'{callee["symbol_name"]} '
                callee_info += f'({callee["file_path"]} {callee["address"]})'
                context_parts.append(callee_info)

        return '\n'.join(context_parts)

    def extract_strings_near_offset(self, elf_file, vaddr, range_size=200):
        """
        提取虚拟地址附近的字符串常量（使用精准提取）

        注意：range_size 参数保留以保持接口兼容性，但实际使用精准提取逻辑
        """
        # 初始化字符串提取器（如果还没有初始化）
        self._init_string_extractor()

        # 使用通用的字符串提取器
        return self.string_extractor.extract_strings_near_offset(elf_file, vaddr)

    def analyze_function(self, file_path, address, call_count, rank, event_count=None, skip_llm=False):
        """分析单个函数（优先使用 radare2，如果不可用则使用 capstone）

        Args:
            file_path: 文件路径
            address: 地址
            call_count: 调用次数
            rank: 排名
            event_count: 事件计数（可选，如果提供则优先显示）
        """
        logger.info(f'\n{"=" * 80}')
        logger.info(f'Analyzing function #{rank}: {file_path}')
        if event_count is not None and event_count > 0:
            logger.info(
                f'Address: {address}, instruction count (event_count): {event_count:,}, call count: {call_count:,}'
            )
        else:
            logger.info(f'Address: {address}, call count: {call_count:,}')
        logger.info(f'{"=" * 80}')

        # 提取虚拟地址（偏移量）
        vaddr = self.extract_offset_from_address(address)
        if vaddr is None:
            logger.warning(' Unable to parse address: %s', address)
            return None

        # 找到 SO 文件
        # 如果指定了 --so-file，且 address 中包含 SO 文件名，优先使用 address 中的信息
        so_file = None
        if self.so_dir and self.so_dir.is_file() and '+' in address:
            # 从 address 中提取 SO 文件名（如 libquickjs.so+0x123 -> libquickjs.so）
            address_so_name = address.split('+')[0]
            # 如果 address 中的 SO 文件名与指定的 SO 文件匹配，直接使用指定的文件
            if address_so_name == self.so_dir.name:
                so_file = self.so_dir.resolve()

        # 如果还没有找到，尝试从 file_path 查找
        if not so_file:
            so_file = self.find_so_file(file_path)

        if not so_file:
            logger.warning(' SO file not found: %s (address: %s)', file_path, address)
            return None

        logger.info(f'✅ Found SO file: {so_file}')

        # 优先使用 radare2（如果可用且未强制使用 Capstone）
        if R2_AVAILABLE and not self.use_capstone_only:
            # 延迟检测 radare2 是否真的可用（只在第一次使用时检测）
            if self._r2_actually_available is None:
                self._r2_actually_available = _check_r2_actually_available()
                if not self._r2_actually_available:
                    logger.info('ℹ️  radare2 command not in PATH, will use capstone for disassembly')

            # 如果 radare2 可用，尝试使用它
            if self._r2_actually_available:
                try:
                    return self._analyze_function_with_r2(
                        so_file,
                        vaddr,
                        file_path,
                        address,
                        call_count,
                        rank,
                        event_count,
                        skip_llm,
                    )
                except (FileNotFoundError, Exception) as e:
                    # 如果 radare2 不可用（FileNotFoundError: ERROR: Cannot find radare2 in PATH）
                    # 标记为不可用，后续直接使用 capstone
                    if isinstance(e, FileNotFoundError) and 'radare2' in str(e):
                        logger.warning(' radare2 command unavailable, automatically switching to capstone method')
                        self._r2_actually_available = False  # 标记为不可用，避免重复尝试
                    else:
                        logger.exception(' radare2 分析失败，回退到 capstone 方法')

        # 使用 capstone 方法（回退方案或默认方案）
        return self._analyze_function_with_capstone(
            so_file, vaddr, file_path, address, call_count, rank, event_count, skip_llm
        )

    def _analyze_function_with_r2(
        self,
        so_file,
        vaddr,
        file_path,
        address,
        call_count,
        rank,
        event_count=None,
        skip_llm=False,
    ):
        """使用 radare2 分析函数（优化：复用同一个 SO 文件的 radare2 实例）"""
        logger.info('🔧 Using radare2 for function analysis')

        # 优化：复用同一个 SO 文件的 radare2 实例，避免重复运行 aaa
        # 确保路径解析一致：统一转换为绝对路径并标准化
        so_file_path_obj = Path(so_file).resolve() if isinstance(so_file, str) else so_file.resolve()

        # 使用标准化的绝对路径作为缓存键（确保路径格式一致）
        # Windows 上路径不区分大小写，统一转换为小写并标准化路径分隔符
        if os.name == 'nt':
            # Windows: 转换为小写，统一使用正斜杠，去除尾部斜杠
            so_file_path = str(so_file_path_obj).lower().replace('\\', '/').rstrip('/')
        else:
            # Unix/Linux: 保持原样，但标准化路径
            so_file_path = str(so_file_path_obj)

        # 确保 _r2_analyzers 已初始化
        if not hasattr(self, '_r2_analyzers'):
            self._r2_analyzers = {}

        # 检查缓存（使用路径对象比较，更可靠）
        cache_match = False
        matched_key = None
        for cached_key in self._r2_analyzers:
            try:
                # 将缓存键和当前路径都转换为 Path 对象并解析为绝对路径进行比较
                cached_path = Path(cached_key).resolve()
                current_path = so_file_path_obj.resolve()
                # 使用 Path 对象的比较（更可靠）
                if cached_path == current_path:
                    cache_match = True
                    matched_key = cached_key
                    break
            except Exception:
                # 如果路径解析失败，回退到字符串比较
                if os.name == 'nt':
                    cached_normalized = cached_key.lower().replace('\\', '/').rstrip('/')
                    current_normalized = so_file_path.lower().replace('\\', '/').rstrip('/')
                else:
                    cached_normalized = cached_key
                    current_normalized = so_file_path
                if cached_normalized == current_normalized:
                    cache_match = True
                    matched_key = cached_key
                    break

        if not cache_match:
            # 第一次打开该 SO 文件，创建并缓存分析器实例
            logger.info(f'📂 First time opening SO file, initializing radare2 analyzer: {so_file_path_obj.name}')
            r2_analyzer = R2FunctionAnalyzer(so_file_path_obj, skip_decompilation=self.skip_decompilation)
            r2_analyzer.__enter__()  # 手动进入上下文管理器，但不退出
            self._r2_analyzers[so_file_path] = r2_analyzer
        else:
            # 复用已存在的分析器实例
            r2_analyzer = self._r2_analyzers[matched_key]
            logger.info(f'♻️  Reusing radare2 analyzer instance: {so_file_path_obj.name} (cached)')

        # 使用缓存的分析器实例进行分析
        result = r2_analyzer.analyze_function_at_offset(vaddr)
        if not result:
            return None

        func_info = result['func_info']
        instructions_str = result['instructions']
        strings = result['strings']
        called_functions = result.get('called_functions', [])  # 获取被调用的函数列表
        decompiled = result.get('decompiled')  # 获取反编译代码（如果可用）

        # 转换指令格式（从字符串转换为列表，保持兼容性）
        instructions = instructions_str  # 已经是字符串列表格式

        # 获取调用堆栈信息（从 perf.db）
        call_stack_info = None
        if self.perf_db_file and self.perf_db_file.exists():
            try:
                call_stack_info = self._get_call_stack_info(file_path, address, vaddr)
            except Exception as e:
                logger.warning(f' Failed to get call stack information: {e}')

        # LLM 分析（如果 skip_llm=True，则跳过）
        llm_result = None
        if self.use_llm and self.llm_analyzer and not skip_llm:
            logger.info('Analyzing function with LLM...')
            try:
                # 构建上下文信息（不包含调用堆栈信息，调用堆栈信息应该在每个函数的上下文中单独显示）
                base_context = self.context if self.context else self._build_context(so_file, file_path)
                context = base_context

                # 获取函数名（如果有）
                symbol_name = func_info.get('name', '')
                if symbol_name.startswith('sym.imp.') or symbol_name.startswith('fcn.'):
                    symbol_name = None  # 这些是导入函数或自动生成的名称，不使用

                llm_result = self.llm_analyzer.analyze_with_llm(
                    instructions=instructions,
                    strings=strings,
                    symbol_name=symbol_name,
                    called_functions=called_functions,  # 传递被调用的函数列表
                    offset=vaddr,
                    context=context,
                    func_info=func_info,  # 传递函数元信息
                    call_count=call_count,  # 传递调用次数
                    event_count=event_count,  # 传递指令执行次数
                    so_file=str(so_file_path) if so_file_path else None,  # 传递 SO 文件路径
                )

                logger.info('✅ LLM analysis completed')
                logger.info(f'   Inferred function name: {llm_result.get("function_name", "N/A")}')
                logger.info(f'   Functionality description: {llm_result.get("functionality", "N/A")[:100]}...')
                logger.info(f'   Confidence: {llm_result.get("confidence", "N/A")}')
            except Exception:
                logger.exception('❌ LLM 分析失败')

        # 返回结果（保持与 capstone 方法相同的格式）
        return {
            'rank': rank,
            'file_path': file_path,
            'address': address,
            'offset': f'0x{vaddr:x}',
            'call_count': call_count,
            'event_count': event_count,  # 添加 event_count
            'so_file': str(so_file),
            'instruction_count': len(instructions),
            'instructions': instructions,  # 字符串列表格式
            'strings': ', '.join(strings[:5]) if strings else '',
            'called_functions': called_functions,  # 添加被调用的函数列表
            'decompiled': decompiled,  # 添加反编译代码（如果可用）
            'llm_result': llm_result,
            'func_info': func_info,  # 额外的函数信息
            'call_stack_info': call_stack_info,  # 添加调用堆栈信息
        }

    def _analyze_function_with_capstone(
        self,
        so_file,
        vaddr,
        file_path,
        address,
        call_count,
        rank,
        event_count=None,
        skip_llm=False,
    ):
        """使用 capstone 分析函数（原有方法）"""
        logger.info('🔧 Using capstone for function analysis')

        # 打开 ELF 文件
        try:
            with open(so_file, 'rb') as f:
                elf_file = ELFFile(f)

                # 反汇编函数
                logger.info(f'Disassembling (vaddr=0x{vaddr:x})...')
                instructions = self.disassemble_function(elf_file, vaddr, size=2000)

                if not instructions:
                    logger.error('❌ Disassembly failed')
                    return None

                logger.info(f'✅ Disassembly successful, {len(instructions)} instructions')

                # 初始化字符串提取器（如果还没有初始化）
                self._init_string_extractor()

                # 提取字符串（传入指令列表以进行精准分析）
                strings = []
                try:
                    extracted_strings = self.string_extractor.extract_strings_from_instructions(
                        elf_file, instructions, vaddr
                    )
                    if extracted_strings:
                        strings = extracted_strings
                        logger.info(
                            f'找到 {len(strings)} 个字符串常量: {", ".join(strings[:3])}{"..." if len(strings) > 3 else ""}'
                        )
                    else:
                        logger.warning(' No string constants found')
                        # 如果精准提取没有找到，尝试 fallback（从 .rodata 段提取）
                        fallback_strings = self.string_extractor._fallback_extract_strings(elf_file)
                        if fallback_strings:
                            strings = fallback_strings
                            logger.info(
                                f'使用 fallback 方法找到 {len(strings)} 个字符串常量: {", ".join(strings[:3])}{"..." if len(strings) > 3 else ""}'
                            )
                except Exception:
                    logger.exception(' 字符串提取失败')
                    strings = []  # 确保 strings 有默认值

                # LLM 分析（如果 skip_llm=True，则跳过）
                llm_result = None
                if self.use_llm and self.llm_analyzer and not skip_llm:
                    logger.info('Analyzing function with LLM...')
                    try:
                        # 构建上下文信息（如果提供了自定义上下文则使用，否则根据 SO 文件自动推断）
                        context = self.context if self.context else self._build_context(so_file, file_path)

                        llm_result = self.llm_analyzer.analyze_with_llm(
                            instructions=[f'{inst.address:x}: {inst.mnemonic} {inst.op_str}' for inst in instructions],
                            strings=strings,
                            symbol_name=None,
                            called_functions=[],
                            offset=vaddr,
                            context=context,
                            func_info=None,  # capstone 模式下没有 func_info
                            call_count=call_count,
                            event_count=event_count,
                            so_file=str(so_file) if so_file else None,
                        )

                        logger.info('✅ LLM analysis completed')
                        logger.info(f'   Inferred function name: {llm_result.get("function_name", "N/A")}')
                        logger.info(f'   Functionality description: {llm_result.get("functionality", "N/A")[:100]}...')
                        logger.info(f'   Confidence: {llm_result.get("confidence", "N/A")}')
                    except Exception:
                        logger.exception('❌ LLM 分析失败')

                # 返回结果
                # 获取调用堆栈信息（从 perf.db）
                call_stack_info = None
                if self.perf_db_file and self.perf_db_file.exists():
                    try:
                        call_stack_info = self._get_call_stack_info(file_path, address, vaddr)
                    except Exception as e:
                        logger.warning(f' Failed to get call stack information: {e}')

                return {
                    'rank': rank,
                    'file_path': file_path,
                    'address': address,
                    'offset': f'0x{vaddr:x}',
                    'call_count': call_count,
                    'event_count': event_count,  # 添加 event_count
                    'so_file': str(so_file),
                    'instruction_count': len(instructions),
                    'instructions': instructions,  # 保存指令用于 HTML 报告
                    'strings': ', '.join(strings[:5]) if strings else '',
                    'called_functions': [],  # capstone 方法暂不支持获取调用函数列表
                    'llm_result': llm_result,
                    'call_stack_info': call_stack_info,  # 添加调用堆栈信息
                }
        except Exception:
            logger.exception('❌ 分析失败')
            return None

    def _get_missing_symbols_from_perf_db(self, top_n=None):
        """从 perf.db 中提取缺失符号（call_count 模式）"""
        if top_n is None:
            top_n = DEFAULT_TOP_N

        logger.info('=' * 80)
        logger.info('⚡️ Extracting missing symbols from perf.db (call_count mode)')
        logger.info('=' * 80)

        conn = sqlite3.connect(str(self.perf_db_file))
        cursor = conn.cursor()

        try:
            # 1. 加载映射关系
            logger.info('\nStep 1: Loading mappings...')
            cursor.execute('SELECT DISTINCT file_id, path FROM perf_files WHERE path IS NOT NULL')
            file_id_to_path = {row[0]: row[1] for row in cursor.fetchall()}
            logger.info(f'✅ Loaded {len(file_id_to_path):,} file path mappings')

            cursor.execute('SELECT id, data FROM data_dict WHERE data IS NOT NULL')
            name_to_data = {row[0]: row[1] for row in cursor.fetchall()}
            logger.info(f'✅ Loaded {len(name_to_data):,} address data mappings')

            # 2. 查询缺失符号记录并聚合
            logger.info('\nStep 2: Querying missing symbol records and aggregating...')
            cursor.execute("""
                SELECT file_id, name, ip, depth
                FROM perf_callchain
                WHERE symbol_id = -1
            """)

            address_call_counts = defaultdict(int)
            address_info = {}

            batch_size = 100000
            total_rows = 0
            while True:
                rows = cursor.fetchmany(batch_size)
                if not rows:
                    break

                total_rows += len(rows)
                for file_id, name_id, _ip, _depth in rows:
                    file_path = file_id_to_path.get(file_id, '未知文件')
                    address_data = name_to_data.get(name_id)

                    if address_data and file_path != '未知文件':
                        key = (file_path, address_data)
                        address_call_counts[key] += 1

                        if key not in address_info:
                            address_info[key] = {
                                'file_path': file_path,
                                'address': address_data,
                            }

            logger.info(
                f'✅ Processing completed, {total_rows:,} records aggregated into {len(address_call_counts):,} unique addresses'
            )

            # 3. 过滤系统文件，并检测 HAP 地址
            excluded_exact = ['[shmm]', '未知文件', '/bin/devhost.elf']
            excluded_prefixes = ['/system', '/vendor/lib64', '/lib', '/chip_prod']

            filtered_data = []
            hap_addresses = []  # 收集 HAP 地址用于批量解析

            for (file_path, address), call_count in address_call_counts.items():
                if file_path in excluded_exact:
                    continue
                if any(file_path.startswith(prefix) for prefix in excluded_prefixes):
                    continue

                # 检测 HAP 地址
                if HAP_RESOLVER_AVAILABLE and is_hap_address(address):
                    hap_addresses.append(address)
                    logger.debug(f'Detected HAP address: {address}')

                filtered_data.append(
                    {
                        'file_path': file_path,
                        'address': address,
                        'call_count': call_count,
                    }
                )

            # 4. 批量解析 HAP 地址（如果存在）
            hap_resolutions = {}
            if hap_addresses and self.perf_db_file and HAP_RESOLVER_AVAILABLE:
                logger.info(f'\nStep 3: Batch resolving {len(hap_addresses)} HAP addresses...')
                hap_resolutions = resolve_hap_addresses_batch(self.perf_db_file, hap_addresses, quick_mode=True)

                # 更新 filtered_data 中的 HAP 地址信息
                for item in filtered_data:
                    address = item['address']
                    if address in hap_resolutions:
                        resolution = hap_resolutions[address]
                        if resolution.get('resolved') and resolution.get('so_file_path'):
                            # 更新为 SO 文件路径和地址
                            item['file_path'] = resolution['so_file_path']
                            item['address'] = f'{resolution["so_name"]}+0x{resolution["so_offset"]:x}'
                            item['hap_resolution'] = resolution
                            logger.info(f'  ✅ {address} -> {item["address"]}')
                        else:
                            item['hap_resolution'] = resolution
                            logger.warning(f'   {address} SO file not found')

            # 5. 如果指定了 --so-file，过滤地址（只保留指定 SO 文件的地址）
            if self.so_dir and self.so_dir.is_file():
                so_filtered_data = []
                for item in filtered_data:
                    address = item.get('address', '')
                    # 从 address 中提取 SO 文件名（如 libquickjs.so+0x123 -> libquickjs.so）
                    if address and '+' in str(address):
                        address_so_name = str(address).split('+')[0]
                        if address_so_name == self.so_dir.name:
                            so_filtered_data.append(item)
                        else:
                            logger.debug(f'跳过不匹配的地址: {address} (期望: {self.so_dir.name})')
                    elif not address or '+' not in str(address):
                        # 如果地址格式不正确，也跳过
                        logger.debug(f'跳过格式不正确的地址: {address}')
                filtered_data = so_filtered_data
                logger.info(f'✅ 过滤后剩余 {len(filtered_data):,} 个地址（指定 SO 文件: {self.so_dir.name}）')

            # 6. 按调用次数排序，取前 N 个
            sorted_data = sorted(filtered_data, key=lambda x: x['call_count'], reverse=True)[:top_n]
            logger.info(f'✅ After filtering, {len(filtered_data):,} records remaining, selecting top {top_n}')

            # 统计 HAP 地址解析情况
            hap_count = sum(1 for item in sorted_data if 'hap_resolution' in item)
            if hap_count > 0:
                resolved_count = sum(1 for item in sorted_data if item.get('hap_resolution', {}).get('resolved'))
                logger.info(f'📊 HAP address statistics: {hap_count} total, {resolved_count} successfully resolved')

            return sorted_data

        finally:
            conn.close()

    def analyze_top_functions(self, top_n=None):
        """分析前 N 个函数"""
        if top_n is None:
            top_n = DEFAULT_TOP_N
        logger.info('=' * 80)
        logger.info(f'Analyzing top {top_n} functions with missing symbols (by call count)')
        logger.info('=' * 80)

        # 如果提供了 excel_file，优先从 Excel 文件读取（因为 Excel 文件可能已经按 event_count 排序）
        if self.excel_file and self.excel_file.exists():
            df = pd.read_excel(self.excel_file)
            logger.info(f'\nReading Excel file: {len(df)} records')

            # 检查是否有 event_count 列，如果有则按 event_count 排序，否则按调用次数排序
            if '指令数(event_count)' in df.columns:
                # 按 event_count 排序
                top_df = df.nlargest(top_n, '指令数(event_count)')
                logger.info(f'Selecting top {top_n} functions (sorted by event_count)')
            elif '调用次数' in df.columns:
                # 按调用次数排序
                top_df = df.nlargest(top_n, '调用次数')
                logger.info(f'Selecting top {top_n} functions (sorted by call count)')
            else:
                # 如果没有排序列，直接取前 top_n 行（保持原有顺序）
                top_df = df.head(top_n)
                logger.info(f'Selecting top {top_n} functions (keeping original order)')

            # 转换为列表格式
            data_list = []
            logger.info(f'从 Excel 读取 {len(top_df)} 行数据，开始过滤...')
            for _, row in top_df.iterrows():
                # 如果指定了 --so-file，过滤地址（从 Excel 读取时也需要过滤）
                if self.so_dir and self.so_dir.is_file():
                    address = row.get('地址', '')
                    # 从 address 中提取 SO 文件名（如 libquickjs.so+0x123 -> libquickjs.so）
                    if address and '+' in str(address):
                        address_so_name = str(address).split('+')[0]
                        # 如果 address 中的 SO 文件名与指定的 SO 文件不匹配，跳过
                        if address_so_name != self.so_dir.name:
                            logger.warning(f'⚠️  从 Excel 读取时跳过不匹配的地址: {address} (期望: {self.so_dir.name})')
                            continue
                    elif not address or '+' not in str(address):
                        # 如果地址格式不正确，也跳过
                        logger.warning(f'⚠️  从 Excel 读取时跳过格式不正确的地址: {address}')
                        continue
                else:
                    # 如果没有指定 --so-file，记录所有地址
                    address = row.get('地址', '')
                    logger.debug(f'从 Excel 读取地址: {address}')

                # 读取 event_count，处理 NaN 值
                event_count = None
                if '指令数(event_count)' in row:
                    event_count_val = row['指令数(event_count)']
                    # 处理 pandas 的 NaN 值
                    if pd.notna(event_count_val) and event_count_val > 0:
                        event_count = int(event_count_val)
                    else:
                        # 如果 event_count 是 NaN 或 0，记录警告
                        logger.warning(f'⚠️  地址 {row.get("地址", "unknown")} 的 event_count 为 NaN 或 0')

                data_list.append(
                    {
                        'file_path': row['文件路径'],
                        'address': row['地址'],
                        'call_count': row.get('调用次数', 0),  # 使用 get 避免 KeyError
                        'event_count': event_count,  # 如果存在且有效，则使用
                    }
                )
        elif self.perf_db_file:
            # 否则从数据库读取
            data_list = self._get_missing_symbols_from_perf_db(top_n)

        # 分析每个函数
        # 如果使用批量 LLM 分析，先收集所有函数信息，然后批量分析
        # 检查是否是批量分析器
        is_batch_analyzer = (
            self.use_llm
            and self.use_batch_llm
            and self.llm_analyzer
            and isinstance(self.llm_analyzer, BatchLLMFunctionAnalyzer)
        )

        if is_batch_analyzer:
            # 批量分析模式：先收集所有函数信息，然后批量分析
            logger.info(f'\nUsing batch LLM analysis mode (batch_size={self.batch_size})')
            results = []
            functions_data = []  # 用于批量分析的数据

            # 第一步：收集所有函数的反汇编和字符串信息（不进行 LLM 分析）
            logger.info('Step 1: Collecting disassembly and string information for all functions...')
            # 确保 _r2_analyzers 已初始化
            if not hasattr(self, '_r2_analyzers'):
                self._r2_analyzers = {}

            for idx, item in enumerate(data_list, 1):
                file_path = item['file_path']
                address = item['address']
                call_count = item.get('call_count', 0)
                event_count = item.get('event_count', None)
                rank = idx

                # 处理 HAP 地址（如果之前没有解析成功，再次尝试）
                if HAP_RESOLVER_AVAILABLE and is_hap_address(address) and self.perf_db_file:
                    hap_resolution = item.get('hap_resolution')
                    if not hap_resolution or not hap_resolution.get('resolved'):
                        logger.info(f'🔄 Re-resolving HAP address: {address}')
                        resolution = resolve_hap_address_from_perfdb(
                            self.perf_db_file, address, quick_mode=True, so_dir=self.so_dir
                        )
                        if resolution and resolution.get('resolved'):
                            file_path = resolution['so_file_path']
                            address = f'{resolution["so_name"]}+0x{resolution["so_offset"]:x}'
                            logger.info(f'  ✅ Resolution successful: {address}')
                        else:
                            logger.warning(f'   Unable to resolve HAP address, skipping: {address}')
                            continue

                # 只进行反汇编和字符串提取，不进行 LLM 分析
                result = self.analyze_function(file_path, address, call_count, rank, event_count, skip_llm=True)
                if result:
                    results.append(result)
                    # 确保调用堆栈信息被正确传递
                    call_stack_info = result.get('call_stack_info')
                    if call_stack_info:
                        logger.debug(
                            f'函数 #{rank} 已有调用堆栈信息: {len(call_stack_info.get("callers", []))} 个调用者, {len(call_stack_info.get("callees", []))} 个被调用者'
                        )
                    else:
                        logger.debug(f'函数 #{rank} 没有调用堆栈信息，将在批量分析时重新获取')
                    # 准备批量分析的数据
                    instructions = result.get('instructions', [])
                    if isinstance(instructions, list) and len(instructions) > 0:
                        # 如果是字符串列表，直接使用；如果是 Instruction 对象，转换为字符串
                        if isinstance(instructions[0], str):
                            inst_list = instructions
                        else:
                            inst_list = [f'{inst.address:x}: {inst.mnemonic} {inst.op_str}' for inst in instructions]
                    else:
                        inst_list = []
                        # 如果没有指令，记录警告但仍然添加到批量分析中（使用空指令列表）
                        logger.warning(
                            f' Function {rank} (address: {address}) has no instructions, will use empty instruction list for LLM analysis'
                        )

                    # 处理字符串：如果是逗号分隔的字符串，转换为列表
                    strings_value = result.get('strings', '')
                    logger.debug(
                        f'Function #{rank} strings_value from result: {repr(strings_value)[:100]}, type: {type(strings_value)}'
                    )
                    if isinstance(strings_value, str):
                        strings_list = (
                            [s.strip() for s in strings_value.split(',') if s.strip()] if strings_value else []
                        )
                    else:
                        strings_list = strings_value if strings_value else []
                    logger.debug(f'Function #{rank} strings_list after processing: {len(strings_list)} items')

                    # 确保所有函数都被添加到批量分析中，即使指令为空
                    # 注意：offset 字段应该使用原始地址字符串（如 "libxxx.so+0x123456"），这样在 prompt 中能正确显示原始地址
                    # result.get('offset') 返回的是 '0x...' 格式，但我们需要原始地址字符串
                    offset_value = address  # 使用原始地址字符串（如 "libquick.so+0xc8ef4"），而不是解析后的偏移量

                    call_stack_info = result.get('call_stack_info')
                    # 如果没有调用堆栈信息，尝试重新获取
                    if not call_stack_info and self.perf_db_file and self.perf_db_file.exists():
                        try:
                            # 从地址中提取偏移量
                            if '+' in address:
                                offset_str = address.split('+', 1)[1]
                                try:
                                    vaddr = (
                                        int(offset_str, 16) if not offset_str.startswith('0x') else int(offset_str, 16)
                                    )
                                    call_stack_info = self._get_call_stack_info(file_path, address, vaddr)
                                    if call_stack_info:
                                        logger.info(
                                            f'函数 #{rank} 重新获取调用堆栈信息成功: {len(call_stack_info.get("callers", []))} 个调用者, {len(call_stack_info.get("callees", []))} 个被调用者'
                                        )
                                except (ValueError, Exception) as e:
                                    logger.debug(f'函数 #{rank} 无法解析地址 {address}: {e}')
                        except Exception as e:
                            logger.debug(f'函数 #{rank} 获取调用堆栈信息失败: {e}')

                    functions_data.append(
                        {
                            'function_id': f'func_{rank}',
                            'offset': offset_value,  # 使用原始地址字符串，确保 prompt 中显示正确的地址
                            'instructions': inst_list,  # 即使为空也添加
                            'strings': strings_list,
                            'symbol_name': None,
                            'called_functions': result.get('called_functions', []),
                            'decompiled': result.get('decompiled'),  # 添加反编译代码
                            'func_info': result.get('func_info'),  # 添加函数元信息
                            'call_count': result.get('call_count', 0),  # 添加调用次数
                            'event_count': result.get('event_count', 0),  # 添加指令执行次数
                            'so_file': result.get('so_file'),  # 添加 SO 文件路径
                            'rank': rank,
                            'file_path': file_path,
                            'address': address,  # 保留原始地址字符串
                            'call_stack_info': call_stack_info,  # 添加调用堆栈信息
                        }
                    )
                    logger.debug(
                        f'Function #{rank} added to functions_data: strings={len(strings_list)}, called_functions={len(result.get("called_functions", []))}, call_stack_info={call_stack_info is not None}'
                    )
                    logger.debug(
                        f'✅ 函数 #{rank} ({address}) 已添加到 functions_data，指令数: {len(inst_list)}, 反编译: {"有" if result.get("decompiled") else "无"}'
                    )

                    # 每10个函数显示一次进度
                    if len(results) % 10 == 0:
                        logger.info(f'  Progress: {len(results)}/{top_n} function information collection completed')

            # 第二步：批量 LLM 分析（按 SO 文件分组）
            if functions_data and self.llm_analyzer:
                logger.info(f'\nStep 2: Batch LLM analysis of {len(functions_data)} functions...')
                logger.debug(
                    f'functions_data 长度: {len(functions_data)}, llm_analyzer: {self.llm_analyzer is not None}'
                )

                # 按 SO 文件分组
                functions_by_so = {}
                for func_data in functions_data:
                    # 从 func_data 中提取 SO 文件名
                    so_file_path = func_data.get('so_file')
                    if not so_file_path:
                        # 如果没有 so_file，尝试从 file_path 或 address 中提取
                        file_path = func_data.get('file_path', '')
                        address = func_data.get('address', '')
                        # 从地址中提取 SO 文件名（如 libquickjs.so+0x123 -> libquickjs.so）
                        if '+' in address:
                            so_name = address.split('+')[0]
                        elif '/' in file_path:
                            so_name = file_path.split('/')[-1]
                        else:
                            so_name = 'unknown'
                    else:
                        # 从 SO 文件路径中提取文件名
                        so_name = Path(so_file_path).name if so_file_path else 'unknown'

                    if so_name not in functions_by_so:
                        functions_by_so[so_name] = []
                    functions_by_so[so_name].append(func_data)

                logger.info(f'按 SO 文件分组: {len(functions_by_so)} 个不同的 SO 文件')
                for so_name, funcs in functions_by_so.items():
                    logger.info(f'  - {so_name}: {len(funcs)} 个函数')

                # 对每个 SO 文件组分别进行批量分析
                all_batch_results = []
                for so_name, so_functions in functions_by_so.items():
                    logger.info(f'\n分析 {so_name} 中的 {len(so_functions)} 个函数...')

                    # 为这个 SO 文件组构建上下文（包含调用堆栈信息）
                    context = None
                    if so_functions:
                        # 使用该组中第一个函数来构建上下文
                        first_func = so_functions[0]
                        so_file_path = first_func.get('so_file')
                        file_path = first_func.get('file_path', '')
                        address = first_func.get('address', '')

                        if so_file_path:
                            so_file = Path(so_file_path)
                            base_context = self.context if self.context else self._build_context(so_file, file_path)
                        elif file_path:
                            # 如果没有 so_file，尝试从 file_path 查找
                            so_file = self.find_so_file(file_path)
                            if so_file:
                                base_context = self.context if self.context else self._build_context(so_file, file_path)
                            else:
                                base_context = self.context if self.context else ''
                        else:
                            base_context = self.context if self.context else ''

                        # 确保每个函数都有调用堆栈信息（如果缺失则重新获取）
                        # 注意：调用堆栈信息是每个地址特有的，不应该合并到整体背景信息中
                        # 每个函数的调用堆栈信息会在 batch_analyzer 中单独显示
                        for func in so_functions:
                            call_stack_info = func.get('call_stack_info')
                            func_file_path = func.get('file_path', '')
                            func_address = func.get('address', '')

                            if (
                                not call_stack_info
                                and self.perf_db_file
                                and self.perf_db_file.exists()
                                and func_file_path
                                and func_address
                            ):
                                try:
                                    # 从地址中提取偏移量
                                    if '+' in func_address:
                                        offset_str = func_address.split('+', 1)[1]
                                        try:
                                            vaddr = int(offset_str, 16)
                                            call_stack_info = self._get_call_stack_info(
                                                func_file_path, func_address, vaddr
                                            )
                                            if call_stack_info:
                                                func['call_stack_info'] = call_stack_info
                                                logger.debug(
                                                    f'函数 {func.get("function_id", "unknown")} 重新获取调用堆栈信息成功: {len(call_stack_info.get("callers", []))} 个调用者, {len(call_stack_info.get("callees", []))} 个被调用者'
                                                )
                                        except (ValueError, Exception) as e:
                                            logger.debug(f'无法解析地址 {func_address}: {e}')
                                except Exception as e:
                                    logger.debug(f'获取调用堆栈信息失败: {e}')

                        # 使用基础上下文，不添加调用堆栈信息（调用堆栈信息会在每个函数中单独显示）
                        context = base_context

                    # 对该 SO 文件组的函数进行批量分析
                    batch_results = self.llm_analyzer.batch_analyze_functions(so_functions, context=context)
                    all_batch_results.extend(batch_results)

                batch_results = all_batch_results

                # 第三步：将 LLM 分析结果合并到 results 中
                logger.info('\nStep 3: Merging LLM analysis results...')
                batch_results_map = {r.get('function_id', ''): r for r in batch_results}
                logger.info(
                    f'批量分析返回了 {len(batch_results)} 个结果，function_id: {[r.get("function_id") for r in batch_results]}'
                )
                logger.info(f'results contains {len(results)} functions, rank: {[r.get("rank") for r in results]}')

                missing_results = []
                for result in results:
                    func_id = f'func_{result["rank"]}'
                    if func_id in batch_results_map:
                        batch_result = batch_results_map[func_id]
                        # 移除 function_id，保留其他字段
                        llm_result = {k: v for k, v in batch_result.items() if k != 'function_id'}
                        result['llm_result'] = llm_result
                        logger.info(
                            f'✅ 函数 #{result["rank"]} ({func_id}, 地址: {result.get("address", "unknown")}) 成功合并 LLM 结果: {llm_result.get("function_name", "None")}, {llm_result.get("functionality", "未知")}'
                        )
                    else:
                        # 如果批量分析结果中没有该函数，使用默认结果
                        missing_results.append(func_id)
                        result['llm_result'] = {
                            'function_name': None,
                            'functionality': '未知',
                            'confidence': '低',
                            'reasoning': '批量 LLM 分析结果中未找到该函数',
                        }
                        logger.warning(
                            f' 函数 #{result["rank"]} ({func_id}, 地址: {result.get("address", "unknown")}) 在批量分析结果中未找到'
                        )

                if missing_results:
                    logger.warning(
                        f' 以下 {len(missing_results)} 个函数在批量分析结果中未找到: {", ".join(missing_results)}'
                    )
        else:
            # 单个分析模式：逐个分析（原有逻辑）
            results = []
            for item in data_list:
                file_path = item['file_path']
                address = item['address']
                call_count = item.get('call_count', 0)
                event_count = item.get('event_count', None)  # 从 Excel 读取时可能包含 event_count
                rank = len(results) + 1

                # 处理 HAP 地址（如果之前没有解析成功，再次尝试）
                if HAP_RESOLVER_AVAILABLE and is_hap_address(address) and self.perf_db_file:
                    hap_resolution = item.get('hap_resolution')
                    if not hap_resolution or not hap_resolution.get('resolved'):
                        logger.info(f'🔄 Re-resolving HAP address: {address}')
                        resolution = resolve_hap_address_from_perfdb(
                            self.perf_db_file, address, quick_mode=True, so_dir=self.so_dir
                        )
                        if resolution and resolution.get('resolved'):
                            file_path = resolution['so_file_path']
                            address = f'{resolution["so_name"]}+0x{resolution["so_offset"]:x}'
                            logger.info(f'  ✅ Resolution successful: {address}')
                        else:
                            logger.warning(f'   Unable to resolve HAP address, skipping: {address}')
                            continue

                result = self.analyze_function(file_path, address, call_count, rank, event_count)
                if result:
                    results.append(result)

                # 每10个函数显示一次进度
                if len(results) % 10 == 0:
                    logger.info(f'\nProgress: {len(results)}/{top_n} function analysis completed')

        # 分析完成后，保存所有缓存和统计（统一在这里调用，避免重复）
        if self.use_llm and self.llm_analyzer:
            self.llm_analyzer.finalize()

        return results

    def _analyze_top_functions_sequential(self, top_df, top_n):
        """逐个分析前 N 个函数（原有逻辑）"""
        results = []
        for _idx, row in top_df.iterrows():
            file_path = row['文件路径']
            address = row['地址']
            call_count = row.get('调用次数', 0)

            # 读取 event_count，处理 NaN 值
            event_count = None
            if '指令数(event_count)' in row:
                event_count_val = row['指令数(event_count)']
                # 处理 pandas 的 NaN 值
                if pd.notna(event_count_val) and event_count_val > 0:
                    event_count = int(event_count_val)

            rank = len(results) + 1

            result = self.analyze_function(file_path, address, call_count, rank, event_count)
            if result:
                results.append(result)

            # 每10个函数显示一次进度
            if len(results) % 10 == 0:
                logger.info(f'\nProgress: {len(results)}/{top_n} function analysis completed')

        # 清理 radare2 分析器（分析完成后释放资源）
        if hasattr(self, '_r2_analyzers'):
            self.cleanup_r2_analyzers()

        return results

    def cleanup_r2_analyzers(self):
        """清理所有缓存的 radare2 分析器实例"""
        if not hasattr(self, '_r2_analyzers'):
            return
        for so_file_path, r2_analyzer in self._r2_analyzers.items():
            try:
                r2_analyzer.__exit__(None, None, None)  # 手动退出上下文管理器
            except Exception as e:
                logger.info(f' Failed to close radare2 analyzer ({Path(so_file_path).name}): {e}')
        self._r2_analyzers.clear()
        logger.info('✅ Cleaned up all radare2 analyzer instances')

    def save_results(
        self,
        results,
        output_file=None,
        html_file=None,
        time_tracker=None,
        top_n=None,
        output_dir=None,
    ):
        """保存分析结果

        Args:
            results: 分析结果列表
            output_file: 输出文件路径（可选，如果不提供则根据 top_n 自动生成）
            html_file: HTML 报告文件路径（可选，如果不提供则根据 output_file 自动生成）
            time_tracker: 时间跟踪器（可选）
            top_n: 分析的数量（用于生成文件名，如果不提供则从结果数量推断）
            output_dir: 输出目录（可选，默认使用配置中的输出目录）
        """
        output_dir = config.get_output_dir() if output_dir is None else Path(output_dir)
        config.ensure_output_dir(output_dir)

        # 如果没有提供 top_n，从结果数量推断
        if top_n is None:
            top_n = len(results) if results else DEFAULT_TOP_N

        # 检查是否使用 event_count（如果结果中有 event_count 字段）
        use_event_count = any('event_count' in result and result.get('event_count', 0) > 0 for result in results)

        # 准备 Excel 数据
        excel_data = []
        for result in results:
            llm_result = result.get('llm_result', {})
            # 确保字符串常量不为 None，空字符串也要保留
            strings_value = result.get('strings', '')
            if strings_value is None:
                strings_value = ''
            elif isinstance(strings_value, list):
                # 如果是列表，转换为逗号分隔的字符串
                strings_value = ', '.join(str(s) for s in strings_value if s)
            elif isinstance(strings_value, str):
                # 如果已经是字符串，直接使用（可能是逗号分隔的字符串）
                strings_value = strings_value if strings_value else ''
            else:
                strings_value = str(strings_value) if strings_value else ''

            # 处理调用链信息（call_stack_info），合并调用者和被调用者信息
            call_stack_info = result.get('call_stack_info')
            call_stack_str = ''

            # 收集所有调用关系
            callers_list = []
            callees_list = []

            # 从 call_stack_info 获取调用者和被调用者
            if call_stack_info:
                callers = call_stack_info.get('callers', [])
                callees = call_stack_info.get('callees', [])

                if callers:
                    for caller in callers[:5]:  # 最多显示5个调用者
                        caller_name = caller.get('symbol_name', '')
                        caller_addr = caller.get('address', '')
                        if caller_name:
                            callers_list.append(f'{caller_name}({caller_addr})')
                        elif caller_addr:
                            callers_list.append(caller_addr)

                if callees:
                    for callee in callees[:5]:  # 最多显示5个被调用者
                        callee_name = callee.get('symbol_name', '')
                        callee_addr = callee.get('address', '')
                        if callee_name:
                            callees_list.append(f'{callee_name}({callee_addr})')
                        elif callee_addr:
                            callees_list.append(callee_addr)

            # 如果没有从 call_stack_info 获取到被调用者，尝试从 called_functions 获取
            if not callees_list:
                called_functions = result.get('called_functions', [])
                if called_functions:
                    callees_list = called_functions[:10]  # 最多显示10个

            # 格式化调用链信息
            call_stack_parts = []
            if callers_list:
                call_stack_parts.append(f'调用者: {", ".join(callers_list)}')
            if callees_list:
                call_stack_parts.append(f'被调用: {", ".join(callees_list)}')

            if call_stack_parts:
                call_stack_str = ' | '.join(call_stack_parts)

            # 获取函数边界信息（用于地址匹配）
            func_info = result.get('func_info', {})
            func_start = func_info.get('minbound', func_info.get('offset', 0))
            func_size = func_info.get('size', 0)
            func_end = func_info.get('maxbound', func_start + func_size) if func_size > 0 else func_start + 2000

            # 确保 llm_result 存在，如果不存在则使用默认值
            if 'llm_result' not in result or result.get('llm_result') is None:
                result['llm_result'] = {
                    'function_name': None,
                    'functionality': '未知',
                    'confidence': '低',
                    'reasoning': 'LLM 分析结果缺失',
                }

            llm_result = result.get('llm_result', {})

            # 安全地获取 LLM 结果字段，确保 None 值转换为空字符串
            def safe_get_llm_field(field_name, llm_result_dict, default=''):
                value = llm_result_dict.get(field_name) if llm_result_dict else None
                if value is None:
                    return default
                result_value = str(value) if value else default
                # 如果是函数名，添加 "Function: " 前缀
                if field_name == 'function_name' and result_value and result_value != default:
                    result_value = f'Function: {result_value}'
                return result_value

            row_data = {
                '排名': result['rank'],
                '文件路径': result['file_path'],
                '地址': result['address'],
                '偏移量': result['offset'],
                'SO文件': result['so_file'],
                '指令数': result['instruction_count'],
                '字符串常量': strings_value,  # 确保是字符串类型，不是 None
                '调用链信息': call_stack_str,  # 合并了调用者和被调用者信息
                'LLM推断函数名': safe_get_llm_field('function_name', llm_result, ''),
                'LLM功能描述': safe_get_llm_field('functionality', llm_result, '未知'),
                '负载问题识别与优化建议': safe_get_llm_field('performance_analysis', llm_result, ''),
                'LLM置信度': safe_get_llm_field('confidence', llm_result, '低'),
                'LLM推理过程': safe_get_llm_field('reasoning', llm_result, ''),
                '函数起始偏移': f'0x{func_start:x}' if func_start else '',
                '函数大小': func_size if func_size > 0 else '',
                '函数结束偏移': f'0x{func_end:x}' if func_end > func_start else '',
            }

            # 根据统计方式选择显示 event_count 或 call_count
            if use_event_count and 'event_count' in result:
                # 使用 event_count 作为主要指标
                row_data['指令数(event_count)'] = result['event_count']
                # 同时显示 call_count（即使为0也显示，用于对比）
                row_data['调用次数'] = result.get('call_count', 0)
            else:
                # 使用 call_count 作为主要指标
                row_data['调用次数'] = result.get('call_count', 0)
                # 如果有 event_count，也显示（用于对比）
                if 'event_count' in result and result.get('event_count', 0) > 0:
                    row_data['指令数(event_count)'] = result['event_count']

            excel_data.append(row_data)

        # 保存 Excel
        df = pd.DataFrame(excel_data)

        # 确保字符串常量列是字符串类型，避免空字符串被转换为 nan
        if '字符串常量' in df.columns:
            df['字符串常量'] = df['字符串常量'].fillna('').astype(str).replace('nan', '').replace('None', '')

        # 确保调用链信息列是字符串类型
        if '调用链信息' in df.columns:
            df['调用链信息'] = df['调用链信息'].fillna('').astype(str).replace('nan', '').replace('None', '')

        # 确保 LLM 相关列不是 nan（处理 None 值）
        llm_columns = ['LLM推断函数名', 'LLM功能描述', '负载问题识别与优化建议', 'LLM置信度', 'LLM推理过程']
        for col in llm_columns:
            if col in df.columns:
                # 将 None 和 nan 替换为空字符串或默认值
                if col == 'LLM功能描述':
                    default_value = '未知'
                elif col == '负载问题识别与优化建议':
                    default_value = ''
                elif col == 'LLM置信度':
                    default_value = '低'
                else:
                    default_value = ''
                df[col] = (
                    df[col]
                    .fillna(default_value)
                    .astype(str)
                    .replace('nan', default_value)
                    .replace('None', default_value)
                )

        # 确保列的顺序：如果使用 event_count，将 event_count 列放在调用次数之前
        if use_event_count and '指令数(event_count)' in df.columns:
            # 重新排列列的顺序
            cols = list(df.columns)
            if '指令数(event_count)' in cols and '调用次数' in cols:
                # 移除这两个列
                cols.remove('指令数(event_count)')
                cols.remove('调用次数')
                # 找到 '偏移量' 的位置，在其后插入
                if '偏移量' in cols:
                    idx = cols.index('偏移量')
                    cols.insert(idx + 1, '指令数(event_count)')
                    cols.insert(idx + 2, '调用次数')
                # 如果没有偏移量，在地址后插入
                elif '地址' in cols:
                    idx = cols.index('地址')
                    cols.insert(idx + 1, '指令数(event_count)')
                    cols.insert(idx + 2, '调用次数')
                else:
                    cols.insert(0, '指令数(event_count)')
                    cols.insert(1, '调用次数')
                df = df[cols]

        if output_file is None:
            excel_file = output_dir / CALL_COUNT_ANALYSIS_PATTERN.format(n=top_n)
        else:
            excel_file = Path(output_file) if isinstance(output_file, str) else output_file
            if not excel_file.is_absolute():
                output_file_str = str(output_file)
                output_dir_str = str(output_dir)
                excel_file = Path(output_file) if output_dir_str in output_file_str else output_dir / excel_file

        # 使用 openpyxl 设置格式
        if not all([Workbook, Alignment, get_column_letter]):
            raise ImportError('openpyxl 未安装，请执行 pip install openpyxl')

        wb = Workbook()
        ws = wb.active
        ws.title = '缺失符号函数分析'

        headers = list(df.columns)
        ws.append(headers)

        for _, row in df.iterrows():
            ws.append([row[col] for col in headers])

        # 设置列宽
        column_widths = {
            '排名': 8,
            '文件路径': 60,
            '地址': 40,
            '偏移量': 15,
            '调用次数': 12,
            '指令数(event_count)': 18,
            'SO文件': 50,
            '指令数': 10,
            '字符串常量': 40,
            '调用链信息': 60,  # 合并了调用者和被调用者信息
            'LLM推断函数名': 30,
            'LLM功能描述': 80,
            '负载问题识别与优化建议': 100,
            'LLM置信度': 12,
            'LLM推理过程': 80,
            '函数起始偏移': 15,
            '函数大小': 12,
            '函数结束偏移': 15,
        }

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = column_widths.get(header, 15)

            # 设置文本换行
            if header in [
                'LLM功能描述',
                '负载问题识别与优化建议',
                'LLM推理过程',
                '文件路径',
                '字符串常量',
                '调用链信息',
            ]:
                for row_idx in range(2, len(df) + 2):
                    cell = ws[f'{col_letter}{row_idx}']
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        wb.save(excel_file)
        logger.info(f'\n✅ Excel report generated: {excel_file}')

        # 不再生成单独的 HTML 报告文件
        # HTML 报告内容将在 Step 4 中直接嵌入到 hiperf_report.html 中
        logger.info('ℹ️  Skipping HTML report file generation (will be embedded in hiperf_report.html in Step 4)')

        return excel_file
